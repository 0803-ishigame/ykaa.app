import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import streamlit_authenticator as stauth
import yaml
from yaml.loader import SafeLoader
from PIL import Image
import openpyxl
from openpyxl.styles import Font
import datetime
import glob
import os

icon = Image.open("company_icon.png")
home_img = Image.open("home.jpg")
#関数
def creat_excel(df_select, house, name):
    font = Font(size=7)
    list1 = list(df_select["使用箇所"])
    list2 = list(df_select["使用部位"])
    list3 = list(df_select["メーカー"])
    list4 = list(df_select["商品名"])
    list5= list(df_select["型番"])
    list6 = list(df_select["色番号"])
    list7 = list(df_select["備考"])
    list_group = [list1, list2, list3, list4, list5, list6, list7]
    wb = openpyxl.load_workbook('仕上げリストテンプ.xlsx')
    ws = wb.active
    dt_now = datetime.datetime.now()
    list_row = ["A", "D", "H", "L", "S", "V", "AB"]
    for i in range(7):
        for j in range(len(list1)):
            ws[f"{list_row[i]}{j+9}"].value = list_group[i][j]
            ws["D6"].value = house
            ws["AD4"].value = dt_now.strftime('%Y/%m/%d %H:%M:%S')
            ws["AF22"].value = name
    rows = ws["A9": "AJ19"]
    for row in rows:
        values = [cell.coordinate for cell in row]
        for value in values: ws[value].font = font 
    output = BytesIO()
    wb.save(output)
    processed_data = output.getvalue()
    return processed_data
def delete_data():
            data_width = 1000
            df = pd.read_excel("仕上げDB.xlsx")
            house_list = []
            where_list = []
            where2_list = []
            df.index = df.index + 1
            df = df.fillna("ー")
            h_ar = df['使用物件'].unique()
            w_ar = df['使用箇所'].unique()
            w2_ar = df['使用部位'].unique()
            for i in range(len(h_ar)):
                house_list.append(h_ar[i])
            for i in range(len(w_ar)):
                where_list.append(w_ar[i])
            for i in range(len(w2_ar)):
                where2_list.append(w2_ar[i])
            st.title("削除")
            col1, col2, col3, col4 = st.columns(4)
            h_select = col1.selectbox('使用物件', house_list)
            w_select = col2.selectbox('使用箇所', where_list)
            w_2_select = col3.selectbox('使用部位', where2_list)
            select_df = df[(df["使用物件"] == h_select) & (df["使用箇所"] == w_select) & (df["使用部位"] == w_2_select)]
            select_df.index = np.arange(1, len(select_df)+1)
            if select_df.empty == True:st.error("データがありません")
            else : st.dataframe(select_df, width=data_width)

            if st.button("削除"):
                df = df.drop(df[(df["使用物件"] == h_select) & (df["使用箇所"] == w_select) & (df["使用部位"] == w_2_select)].index, inplace=False)
                df.to_excel("仕上げDB.xlsx", index=False)
def add_data():
    df = pd.read_excel("仕上げDB.xlsx")
    house_list = []
    where_list = []
    where2_list = []
    df.index = df.index + 1
    df = df.fillna("ー")
    h_ar = df['使用物件'].unique()
    w_ar = df['使用箇所'].unique()
    w2_ar = df['使用部位'].unique()
    for i in range(len(h_ar)):
        house_list.append(h_ar[i])
    for i in range(len(w_ar)):
        where_list.append(w_ar[i])
    for i in range(len(w2_ar)):
        where2_list.append(w2_ar[i])
    st.title("編集")
    col1, col2, col3, col4 = st.columns(4)
    h_select = col1.selectbox('使用物件', house_list)
    w_select = col2.selectbox('使用箇所', where_list)
    w_2_select = col3.selectbox('使用部位', where2_list)
    df_select = df[(df['使用物件'] == h_select) & (df['使用箇所'] == w_select) & (df['使用部位'] == w_2_select)]
    df = df[(df['使用物件'] != h_select) | (df['使用箇所'] != w_select) | (df['使用部位'] != w_2_select)]
    if df_select.empty == True:st.error("データがありません")
        #after
    else:
        company = st.text_input("メーカー", value=df_select.iat[0, 3])
        goods_name = st.text_input("商品名", value=df_select.iat[0, 4])
        goods_number = st.text_input("型番", value=df_select.iat[0, 5])
        goods_color = st.text_input("色番号", value=df_select.iat[0, 6])
        etc = st.text_input("備考", value=df_select.iat[0, 7])
        df_new = pd.DataFrame({"使用物件": [h_select], "使用箇所":[w_select], "使用部位":[w_2_select], "メーカー":[company], "商品名":[goods_name], "型番":[goods_number], "色番号":[goods_color], "備考":[etc]})
        df_new.index = df_new.index + 1
        if (df_new.iat[0,3] == df_select.iat[0,3])&(df_new.iat[0,4] == df_select.iat[0,4])&(df_new.iat[0,5] == df_select.iat[0,5])&(df_new.iat[0,6] == df_select.iat[0,6])&(df_new.iat[0,7] == df_select.iat[0,7]):    
            st.success("編集中")
        else:
            df = pd.concat([df, df_new])
            df.to_excel("仕上げDB.xlsx",index=False)
            st.success("保存完了")
            df_new = df_select
def new_data():
    df = pd.read_excel("仕上げDB.xlsx")
    st.title("新規作成")
    house = st.text_input("使用物件")
    col1, col2 = st.columns(2)
    where = col1.text_input("使用箇所")
    where2 = col2.text_input("使用部位")
    col_1, col_2, col_3 = st.columns(3)
    company = col_1.text_input("メーカー")
    goods_name = col_2.text_input("商品名")
    goods_namber = col_3.text_input("型番")
    col__1, col__2 = st.columns(2)
    goods_color = col__1.text_input("色番号")
    etc = col__2.text_input("備考")
    if st.button("新規作成"):
        df_new = pd.DataFrame({"使用物件": [house], "使用箇所":[where], "使用部位":[where2], "メーカー":[company], "商品名":[goods_name], "型番":[goods_namber], "色番号":[goods_color], "備考":[etc]})
        df = pd.concat([df, df_new])
        df.to_excel("仕上げDB.xlsx",index=False)
        st.success("成功しました")
st.set_page_config(
    page_title="YKAA",
    layout="wide",
    page_icon=icon,
)

## ユーザー設定読み込み
yaml_path = "config.yaml"

with open(yaml_path) as file:
    config = yaml.load(file, Loader=SafeLoader)

authenticator = stauth.Authenticate(
    credentials=config['credentials'],
    cookie_name=config['cookie']['name'],
    cookie_key=config['cookie']['key'],
    cookie_expiry_days=config['cookie']['expiry_days'],
)

## UI 
authenticator.login()
if st.session_state["authentication_status"]:
    ## ログイン成功
    with st.sidebar:
        name = st.session_state["name"]
        st.image(icon)
        if name == 'Takashi Ishigame': name_ja, job = '石亀　隆', 'スタッフ'
        if name == 'Yoshiki Katayama': name_ja, job = '片山　佳紀', '設計主任'
        if name == 'Hayato Kumagai': name_ja, job = '熊谷　颯人', 'スタッフ'
        if name == 'Yoshiharu Kikuchi': name_ja, job = '菊池　佳晴', '所長'
        if name == 'Yuko Kikuchi': name_ja, job = '菊池　祐子', '取締役'
        if name == 'Masaki Okazaki': name_ja, job = '岡崎　雅樹', '設計主任'
        st.markdown(f'## {job} : {name_ja}')
        authenticator.logout('Logout', 'sidebar')
        st.divider()
    
    menu = ["HOME","使用仕上げ一覧", "仕様建材一覧"]
    state = st.sidebar.selectbox("Application", menu)

    if state == "HOME":
        st.title("株式会社菊池佳晴建築設計事務所スタッフページ")
        st.image(home_img)

    if state == "使用仕上げ一覧":
        
        ##使用仕上げ一覧アプリ
        menu_sub = ["一覧", "新規作成", "編集", "削除"]
        state_sub = st.sidebar.selectbox("menu", menu_sub)
        if state_sub == "一覧":
            data_width = 1500
            house_list = ["指定なし"]
            where_list = ["指定なし"]
            df = pd.read_excel("仕上げDB.xlsx")
            df = df.fillna("ー")
            df.index = df.index + 1
            h_ar = df['使用物件'].unique()
            w_ar = df['使用箇所'].unique()
            for i in range(len(h_ar)):
                house_list.append(h_ar[i])
            for i in range(len(w_ar)):
                where_list.append(w_ar[i])

            st.title("使用仕上げ一覧")
            col1, col2 = st.columns(2)
            h_select = col1.selectbox('使用物件', house_list)
            w_select = col2.selectbox('使用箇所', where_list)
            if h_select == "指定なし" and w_select == "指定なし":
                st.dataframe(df, width=data_width)
                select_df = df
            elif w_select == "指定なし":
                select_df = df[df["使用物件"] == h_select]
                select_df.index = np.arange(1, len(select_df)+1)
                st.dataframe(select_df, width=data_width)
            elif h_select == "指定なし":
                select_df = df[df["使用箇所"] == w_select]
                select_df.index = np.arange(1, len(select_df)+1)
                st.dataframe(select_df, width=data_width)
            else:
                select_df = df[(df["使用物件"] == h_select) & (df["使用箇所"] == w_select)]
                select_df.index = np.arange(1, len(select_df)+1)
                st.dataframe(select_df, width=data_width)
            df_xlsx = creat_excel(select_df, h_select, name_ja)
            if h_select == "指定なし":h=""
            else:h= h_select + "_"
            st.download_button("EXCEL保存", df_xlsx, file_name=f"{h}仕上げリスト.xlsx")
        if state_sub == "新規作成":
            new_data()
        if state_sub == "削除":
            delete_data()
        if state_sub == "編集":
            add_data()

    if state == "仕様建材一覧":
        menu_sub = ["一覧", "新規作成", "編集"]
        state_sub = st.sidebar.selectbox("menu", menu_sub)
        if state_sub == "一覧":
            data_width = 1500
            com_list = ["指定なし"]
            goods_list = ["指定なし"]
            df_list = pd.read_excel('建材DB.xlsx')
            df_list = df_list.fillna("ー")
            com_l = df_list["メーカー"].unique()
            for i in range(len(com_l)):
                com_list.append(com_l[i])
            col_1, col_2, col_3 = st.columns(3)
            c_select = col_1.selectbox("メーカー", com_list)
            df_list_select = df_list[df_list["メーカー"] == c_select]
            goods_l = df_list_select["商品名"].unique()
            for i in range(len(goods_l)):
                goods_list.append(goods_l[i])
            g_select = col_2.selectbox("商品名", goods_list)
            df_list.index = df_list.index + 1
            if c_select == "指定なし" and g_select == "指定なし":
                st.dataframe(df_list, width=data_width)
                select_df = df_list
            elif c_select == "指定なし":
                select_df = df_list[df_list["商品名"] == g_select]
                select_df.index = np.arange(1, len(select_df)+1)
                st.dataframe(select_df, width=data_width)
            elif g_select == "指定なし":
                select_df = df_list[df_list["メーカー"] == c_select]
                select_df.index = np.arange(1, len(select_df)+1)
                st.dataframe(select_df, width=data_width)
            else:
                select_df = df_list[(df_list["メーカー"] == c_select) & (df_list["商品名"] == g_select)]
                select_df.index = np.arange(1, len(select_df)+1)
                st.dataframe(select_df, width=data_width)
        if state_sub == "新規作成":
            com_list = ["新規メーカー"]    
            df_goods = pd.read_excel("建材DB.xlsx")
            df_com = pd.read_excel("メーカーDB.xlsx")
            com_l = df_com["メーカー"].unique()
            for i in range(len(com_l)):
                com_list.append(com_l[i])
            st.title("新規作成")
            col1, col2, col3 = st.columns(3)
            company = col1.selectbox("メーカー", com_list)
            if company == "新規メーカー":
                company = col2.text_input("新規メーカー名")
                dictionary = col2.checkbox("カタログ")
                if dictionary: col3.number_input("発行年数", 2024)
            col_1, col_2, col_3 = st.columns(3)
            goods_name = col_1.text_input("商品名")
            goods_namber = col_2.text_input("型番")
            goods_color = col_3.text_input("色番号")
            col__1, col__2 = st.columns(2)
            sample = col__1.checkbox("サンプル")
            if sample:
                sample_state = '◯'
                sample_photo_path = st.file_uploader("サンプル写真", type=['.jpeg', '.png'], accept_multiple_files=False)
                if sample_photo_path != None:
                    sample_img = Image.open(sample_photo_path)
            else: sample_state = '☓'
            
            etc = st.text_input("備考")
            if st.button("新規作成"):
                if "photo" not in [f for f in os.listdir("./photo") if os.path.isdir(os.path.join("./photo", f))]:
                    os.mkdir(f"./photo")
                if company not in [f for f in os.listdir("./photo") if os.path.isdir(os.path.join("./photo", f))]:
                    os.mkdir(f"./photo/{company}")
                if sample_img != None:
                    file_name = f"{goods_name}_{goods_namber}_{goods_color}.jpeg"
                    sample_img.save(f"./photo/{company}/{file_name}")
                df_new = pd.DataFrame({"メーカー":[company], "商品名":[goods_name], "型番":[goods_namber], "色番号":[goods_color], "サンプル":[sample_state] ,"備考":[etc]})
                df_new = df_new.fillna("ー")
                df_goods = pd.concat([df_goods, df_new])
                df_goods.to_excel("建材DB.xlsx",index=False)
                result = Image.open(f"./photo/{company}/{file_name}")
                st.image(result, width=200)
                st.success("成功しました")



elif st.session_state["authentication_status"] is False:
    ## ログイン成功ログイン失敗
    st.error('Username/password is incorrect')




